# _*_ coding:utf-8 _*_
# 作者：Olivia
# @Time: 2021/3/2 下午11:02
# @File: test_excel_dataDriver.py

'''
excel数据驱动
1. excel操作流程：先操作工作簿 wookbook，再操作sheet页，再操作单元格 cell(value)
'''

import openpyxl
# 操作工作簿：指定文件路径，进行文件读取
excel = openpyxl.load_workbook('../data/caseData.xlsx')
# 或许sheet，基于工作簿来获取sheet页
sheetNames = excel.sheetnames
print(sheetNames)
for name in sheetNames:
    print(name)
# 操作单元格cell
rootCreateParkSheet = excel['rootCreatePark']
print(rootCreateParkSheet)
# 获取单元格内容
print(rootCreateParkSheet.values)
for value in rootCreateParkSheet.values:
    print(value)

sheet1 = excel['Sheet1']
print(sheet1)
# 获取单元格内容
print(sheet1.values)
for value in sheet1.values:
    print(value[0])

# 获取最大行数
row = sheet1.max_row
print(row)

# 获取最大列数
column = sheet1.max_column
print(column)

# 修改单元格内容
sheet1['A3'] = 'A'
# 保存
excel.save('../data/caseData.xlsx')
# 释放
excel.close()
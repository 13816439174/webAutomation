# _*_ coding:utf-8 _*_
# 作者：Olivia
# @Time: 2021/3/8 下午8:30
# @File: test_06_excelAndKeywordAndUnittestQmailLogin.py

import unittest
from base.log import log_option
import openpyxl
from base.base_keyword import WebKeys
from openpyxl.styles import Font,PatternFill
from base.read_excel import read_excel

class Demo(unittest.TestCase):

    @classmethod
    def setUpClass(cls)->None:
        print('this is setUpClass')
        cls.log = log_option()
        # cls.wb = WebKeys()

    # def tearDownClass(cls)->None:
    #     print('this is tearDownclass')
    @unittest.skip('skip')
    def test_01_qqMailLoginFailed(self):
        excel=openpyxl.load_workbook('../data/qqEmailLogin.xlsx')
        LoginFailedSheet = excel['LoginFailed']
        self.log.info('获取{0}内容成功，现在开始自动化测试>>>>>'.format(LoginFailedSheet))
        for value in LoginFailedSheet.values:
            args={}
            args['name']=value[2]
            args['value']=value[3]
            args['txt']=value[4]
            args['expect']=value[6]
            # 基于A列进行判断是否为测试用例
            if type(value[0]) is int:
                # 判断是否实例化
                if value[1] == 'open_browser':
                    self.log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
                    wk=WebKeys(value[4])
                # 断言关键字执行后：首先执行断言，再判断断言是否成功
                elif 'assert' in value[1]:
                    self.log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
                    # 依据status来判断写入的是Failed，还是Pass
                    status=getattr(wk, value[1])(**args)
                    if status:
                        # 写入当前行的实际结果值
                        # sheet.cell(row='编号+1',column=8)
                        LoginFailedSheet.cell(row=value[0] + 1, column=8).value='Pass'
                        LoginFailedSheet.cell(row=value[0] + 1, column=8).fill=PatternFill('solid', fgColor='AACF91')
                        LoginFailedSheet.cell(row=value[0] + 1, column=8).font=Font(bold=True)
                    else:
                        LoginFailedSheet.cell(row=value[0] + 1, column=8).value='Failed'
                        LoginFailedSheet.cell(row=value[0] + 1, column=8).fill=PatternFill('solid', fgColor='FF0000')
                        LoginFailedSheet.cell(row=value[0] + 1, column=8).font=Font(bold=True)
                    excel.save('../data/qqEmailLogin.xlsx')
                # 非特殊关键字，通过反射机制实现
                else:
                    self.log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
                    getattr(wk, value[1])(**args)
        excel.close()

    def test_02(self):
        read_excel('../data/qqEmailLogin.xlsx')



if __name__ == '__main__':
    unittest.main()

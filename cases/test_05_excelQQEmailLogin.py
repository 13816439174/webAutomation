# _*_ coding:utf-8 _*_
# 作者：Olivia
# @Time: 2021/3/4 下午2:37
# @File: test_05_excelQQEmailLogin.py

import openpyxl
from base.base_keyword import WebKeys
from base.log import log_option
from openpyxl.styles import Font,PatternFill

log = log_option()
excel = openpyxl.load_workbook('../data/qqEmailLogin.xlsx')
sheetNames = excel.sheetnames
print(sheetNames)
for name in sheetNames:
    print(name)

LoginFailedSheet = excel['LoginFailed']
print(LoginFailedSheet.values)

log.info('获取{0}内容成功，现在开始自动化测试>>>>>'.format(LoginFailedSheet))
for value in LoginFailedSheet.values:
    # print(value)
    # 定义一个字典参数，用于接收excel中的所有参数内容
    args = {}
    args['name'] = value[2]
    args['value'] = value[3]
    args['txt'] = value[4]
    args['expect'] = value[6]
    # 基于A列进行判断是否为测试用例
    if type(value[0]) is int:
        '''
        在读取关键字时，分几类情况：
        1。 关键字驱动类的实例化
        2。 断言类型的关键字：比较特殊
            因为断言的关键字除去，预期结果和实际结果进行对比。还要将实际结果填写在excel中
        '''
        # 判断是否实例化
        if value[1] == 'open_browser':
            log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
            wk = WebKeys(value[4])
        # 断言关键字执行后：首先执行断言，再判断断言是否成功
        elif 'assert' in value[1]:
            log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
            # 依据status来判断写入的是Failed，还是Pass
            status = getattr(wk,value[1])(**args)
            if status:
                # 写入当前行的实际结果值
                # sheet.cell(row='编号+1',column=8)
                LoginFailedSheet.cell(row=value[0]+1, column=8).value = 'Pass'
                LoginFailedSheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid',fgColor='AACF91')
                LoginFailedSheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            else:
                LoginFailedSheet.cell(row=value[0]+1,column=8).value = 'Failed'
                LoginFailedSheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='FF0000')
                LoginFailedSheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            excel.save('../data/qqEmailLogin.xlsx')
        # 非特殊关键字，通过反射机制实现
        else:
            log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
            getattr(wk,value[1])(**args)

# excel.save('../data/qqEmailLogin.xlsx')
excel.close()
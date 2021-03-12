# _*_ coding:utf-8 _*_
# 作者：Olivia
# @Time: 2021/3/11 下午8:30
# @File: read_excel.py
import os
import openpyxl
from base.log import log_option
from openpyxl.styles import PatternFill,Font
from base.base_keyword import WebKeys

def get_excel(path):
    a=os.path.exists(path)
    if a:
        excel = openpyxl.load_workbook(path)
        return excel
    else:
        return False

def get_excelSheet(path):
    a=os.path.exists(path)
    if a:
        excel=openpyxl.load_workbook(path)
        sheetNames=excel.sheetnames
        name = sheetNames[0]
        sheet=excel[name]
        # print(sheet)
        return sheet
        # for value in sheet.values:
        #     print(value)
    else:
        print('path is not exit')
        return False

def read_excel(path):
    log = log_option()
    excel = get_excel(path)
    sheet = get_excelSheet(path)
    for value in sheet.values:
        args={}
        args['name']=value[2]
        args['value']=value[3]
        args['txt']=value[4]
        args['expect']=value[6]
        # 基于A列进行判断是否为测试用例
        if type(value[0]) is int:
            # 判断是否实例化
            if value[1] == 'open_browser':
                log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
                wk=WebKeys(value[4])
            # 断言关键字执行后：首先执行断言，再判断断言是否成功
            elif 'assert' in value[1]:
                log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
                # 依据status来判断写入的是Failed，还是Pass
                status=getattr(wk, value[1])(**args)
                if status:
                    # 写入当前行的实际结果值
                    # sheet.cell(row='编号+1',column=8)
                    sheet.cell(row=value[0] + 1, column=8).value='Pass'
                    sheet.cell(row=value[0] + 1, column=8).fill=PatternFill('solid', fgColor='AACF91')
                    sheet.cell(row=value[0] + 1, column=8).font=Font(bold=True)
                else:
                    sheet.cell(row=value[0] + 1, column=8).value='Failed'
                    sheet.cell(row=value[0] + 1, column=8).fill=PatternFill('solid', fgColor='FF0000')
                    sheet.cell(row=value[0] + 1, column=8).font=Font(bold=True)
                excel.save(path)
            # 非特殊关键字，通过反射机制实现
            else:
                log.info('现在执行关键字:{0}，操作描述:{1}'.format(value[1], value[5]))
                getattr(wk, value[1])(**args)




# get_excelSheet('../data/qqEmailLogin.xlsx')
# read_excel('../data/qqEmailLogin.xlsx')

